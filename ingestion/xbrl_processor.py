"""
xbrl_processor.py
=================
Processes XBRL financial data from XML filings and Excel workbooks,
inserting structured facts into the PostgreSQL xbrl_facts table.

Features:
  - XBRLFact dataclass for structured fact representation
  - Parses XBRL/XML filings using lxml (IFRS taxonomy)
  - Parses Excel workbooks with XBRL-like data using openpyxl
  - SHA-256 content hash for deduplication
  - Batch insert with ON CONFLICT DO NOTHING on content_hash
  - Skip already-processed filings
  - process_filing(file_path), process_directory(dir_path), process_url(url)

Usage:
    # Process a single XML filing
    python ingestion/xbrl_processor.py --file data/filing.xml --ticker 2222.SR

    # Process a single workbook
    python ingestion/xbrl_processor.py --file data/filing.xlsx --ticker 2222.SR

    # Process a directory of filings
    python ingestion/xbrl_processor.py --dir data/filings/ --ticker-pattern "*.xml"

    # Dry run
    python ingestion/xbrl_processor.py --file data/filing.xml --ticker 2222.SR --dry-run
"""

import argparse
import hashlib
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

try:
    from lxml import etree
except ImportError:
    etree = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None

try:
    import requests
except ImportError:
    requests = None

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
BATCH_SIZE = 250

# XBRL namespace prefixes used in Saudi IFRS filings
XBRL_NAMESPACES = {
    "xbrli": "http://www.xbrl.org/2003/instance",
    "ifrs-full": "http://xbrl.ifrs.org/taxonomy/2023-03-23/ifrs-full",
    "link": "http://www.xbrl.org/2003/linkbase",
    "xlink": "http://www.w3.org/1999/xlink",
    "iso4217": "http://www.xbrl.org/2003/iso4217",
    "xbrldi": "http://xbrl.org/2006/xbrldi",
}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class XBRLFact:
    """A single XBRL fact extracted from a financial filing."""
    ticker: str
    concept: str                          # XBRL concept (e.g., 'ifrs-full:Revenue')
    label_en: Optional[str] = None
    label_ar: Optional[str] = None
    value_numeric: Optional[float] = None
    value_text: Optional[str] = None
    value_boolean: Optional[bool] = None
    unit: Optional[str] = None            # e.g., 'SAR', 'shares', 'pure'
    decimals: Optional[int] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    period_instant: Optional[date] = None
    dimension_member: Optional[str] = None
    dimension_value: Optional[str] = None
    source_url: Optional[str] = None
    filing_id: Optional[str] = None       # UUID string of the parent filing
    content_hash: str = field(default="", init=False)

    def __post_init__(self):
        """Compute content hash for deduplication."""
        self.content_hash = self._compute_hash()

    def _compute_hash(self) -> str:
        """SHA-256 hash of the fact's identity fields for dedup."""
        parts = [
            self.ticker,
            self.concept,
            str(self.value_numeric) if self.value_numeric is not None else "",
            self.value_text or "",
            str(self.value_boolean) if self.value_boolean is not None else "",
            self.unit or "",
            str(self.period_start) if self.period_start else "",
            str(self.period_end) if self.period_end else "",
            str(self.period_instant) if self.period_instant else "",
            self.dimension_member or "",
            self.dimension_value or "",
        ]
        raw = "|".join(parts)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def to_insert_tuple(self) -> tuple:
        """Return a tuple matching the xbrl_facts INSERT column order."""
        return (
            self.ticker,
            self.filing_id,
            self.concept,
            self.label_en,
            self.label_ar,
            self.value_numeric,
            self.value_text,
            self.value_boolean,
            self.unit,
            self.decimals,
            self.period_start,
            self.period_end,
            self.period_instant,
            self.dimension_member,
            self.dimension_value,
            self.source_url,
            self.content_hash,
        )


# Column order for INSERT
XBRL_INSERT_COLUMNS = [
    "ticker", "filing_id", "concept", "label_en", "label_ar",
    "value_numeric", "value_text", "value_boolean",
    "unit", "decimals",
    "period_start", "period_end", "period_instant",
    "dimension_member", "dimension_value",
    "source_url", "content_hash",
]

XBRL_INSERT_SQL = (
    f"INSERT INTO xbrl_facts ({', '.join(XBRL_INSERT_COLUMNS)}) "
    f"VALUES ({', '.join(['%s'] * len(XBRL_INSERT_COLUMNS))}) "
    f"ON CONFLICT (content_hash) DO NOTHING"
)


# ---------------------------------------------------------------------------
# XBRLProcessor
# ---------------------------------------------------------------------------

class XBRLProcessor:
    """Parses XBRL XML filings and Excel workbooks, extracting financial facts."""

    # Common IFRS concept mappings for label-to-concept conversion
    IFRS_CONCEPT_MAP = {
        "total assets": "ifrs-full:Assets",
        "total liabilities": "ifrs-full:Liabilities",
        "total equity": "ifrs-full:Equity",
        "revenue": "ifrs-full:Revenue",
        "total revenue": "ifrs-full:Revenue",
        "cost of revenue": "ifrs-full:CostOfSales",
        "cost of sales": "ifrs-full:CostOfSales",
        "gross profit": "ifrs-full:GrossProfit",
        "operating income": "ifrs-full:ProfitLossFromOperatingActivities",
        "net income": "ifrs-full:ProfitLoss",
        "profit for the period": "ifrs-full:ProfitLoss",
        "earnings per share": "ifrs-full:BasicEarningsLossPerShare",
        "basic eps": "ifrs-full:BasicEarningsLossPerShare",
        "diluted eps": "ifrs-full:DilutedEarningsLossPerShare",
        "cash and cash equivalents": "ifrs-full:CashAndCashEquivalents",
        "property plant and equipment": "ifrs-full:PropertyPlantAndEquipment",
        "retained earnings": "ifrs-full:RetainedEarnings",
        "share capital": "ifrs-full:IssuedCapital",
        "operating cash flow": "ifrs-full:CashFlowsFromOperatingActivities",
        "investing cash flow": "ifrs-full:CashFlowsFromInvestingActivities",
        "financing cash flow": "ifrs-full:CashFlowsFromFinancingActivities",
        "dividends paid": "ifrs-full:DividendsPaid",
        "depreciation": "ifrs-full:DepreciationAndAmortisationExpense",
        "interest expense": "ifrs-full:InterestExpense",
        "interest income": "ifrs-full:InterestRevenueForFinancialAssetsNotAtFairValue",
        "income tax": "ifrs-full:IncomeTaxExpenseContinuingOperations",
        "current assets": "ifrs-full:CurrentAssets",
        "non-current assets": "ifrs-full:NoncurrentAssets",
        "current liabilities": "ifrs-full:CurrentLiabilities",
        "non-current liabilities": "ifrs-full:NoncurrentLiabilities",
        "accounts receivable": "ifrs-full:TradeAndOtherCurrentReceivables",
        "inventory": "ifrs-full:Inventories",
        "accounts payable": "ifrs-full:TradeAndOtherCurrentPayables",
        "long-term debt": "ifrs-full:NoncurrentBorrowings",
        "short-term debt": "ifrs-full:CurrentBorrowings",
    }

    # Common sheet names in Saudi XBRL Excel exports
    EXPECTED_SHEETS = [
        "Balance Sheet", "Statement of Financial Position",
        "Income Statement", "Statement of Profit or Loss",
        "Cash Flow", "Statement of Cash Flows",
        "Changes in Equity", "Notes",
    ]

    def __init__(
        self,
        ticker: str,
        filing_id: Optional[str] = None,
        source_url: Optional[str] = None,
        default_unit: str = "SAR",
    ):
        self.ticker = ticker
        self.filing_id = filing_id
        self.source_url = source_url
        self.default_unit = default_unit
        self.facts: list[XBRLFact] = []
        self.errors: list[str] = []

    # ------------------------------------------------------------------
    # Public API: process_filing, process_directory, process_url
    # ------------------------------------------------------------------

    def process_filing(self, file_path: Path) -> list[XBRLFact]:
        """Process a single filing (XML or Excel) and return extracted facts.

        Dispatches to XML or Excel parser based on file extension.

        Args:
            file_path: Path to the filing file (.xml, .xbrl, .xlsx, .xls).

        Returns:
            List of XBRLFact objects extracted from the filing.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            self.errors.append(f"File not found: {file_path}")
            return []

        suffix = file_path.suffix.lower()
        if suffix in (".xml", ".xbrl"):
            return self.process_xml(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self.process_workbook(file_path)
        else:
            self.errors.append(f"Unsupported file type: {suffix}")
            return []

    def process_directory(self, dir_path: Path, pattern: str = "*") -> list[XBRLFact]:
        """Process all filing files in a directory.

        Args:
            dir_path: Directory containing filing files.
            pattern: Glob pattern for matching files (default: all files).

        Returns:
            Combined list of XBRLFact objects from all files.
        """
        dir_path = Path(dir_path)
        if not dir_path.exists():
            self.errors.append(f"Directory not found: {dir_path}")
            return []

        all_facts = []
        supported_extensions = {".xml", ".xbrl", ".xlsx", ".xls"}
        files = sorted(dir_path.glob(pattern))

        for file_path in files:
            if file_path.suffix.lower() not in supported_extensions:
                continue
            try:
                facts = self.process_filing(file_path)
                all_facts.extend(facts)
                logger.info(
                    "Processed %s: %d facts extracted", file_path.name, len(facts)
                )
            except Exception as e:
                self.errors.append(f"Error processing {file_path.name}: {e}")
                logger.error("Error processing %s: %s", file_path.name, e)

        return all_facts

    def process_url(self, url: str, download_dir: Optional[Path] = None) -> list[XBRLFact]:
        """Download a filing from a URL and process it.

        Args:
            url: URL of the filing to download.
            download_dir: Directory to save the downloaded file.
                          Defaults to a temp location under SCRIPT_DIR.

        Returns:
            List of XBRLFact objects extracted from the filing.
        """
        if requests is None:
            self.errors.append("requests library required for URL processing: pip install requests")
            return []

        self.source_url = url
        if download_dir is None:
            download_dir = SCRIPT_DIR / "_downloads"
        download_dir = Path(download_dir)
        download_dir.mkdir(parents=True, exist_ok=True)

        # Derive filename from URL
        parsed = urlparse(url)
        filename = Path(parsed.path).name or "filing.xml"
        local_path = download_dir / filename

        try:
            logger.info("Downloading %s", url)
            resp = requests.get(url, timeout=60)
            resp.raise_for_status()
            local_path.write_bytes(resp.content)
            logger.info("Saved to %s (%d bytes)", local_path, len(resp.content))
        except Exception as e:
            self.errors.append(f"Failed to download {url}: {e}")
            return []

        return self.process_filing(local_path)

    # ------------------------------------------------------------------
    # XML/XBRL parsing (lxml)
    # ------------------------------------------------------------------

    def process_xml(self, file_path: Path) -> list[XBRLFact]:
        """Parse an XBRL XML filing and extract facts.

        Handles IFRS taxonomy elements. Each fact element in the XBRL instance
        document is mapped to an XBRLFact.

        Args:
            file_path: Path to the XML/XBRL file.

        Returns:
            List of XBRLFact objects.
        """
        if etree is None:
            raise ImportError("lxml is required for XML parsing: pip install lxml")

        self.facts = []
        try:
            tree = etree.parse(str(file_path))
        except etree.XMLSyntaxError as e:
            self.errors.append(f"XML syntax error in {file_path}: {e}")
            return []
        except Exception as e:
            self.errors.append(f"Cannot parse XML {file_path}: {e}")
            return []

        root = tree.getroot()
        nsmap = root.nsmap.copy()
        # Ensure default namespace doesn't break XPath
        if None in nsmap:
            nsmap["default"] = nsmap.pop(None)

        # Parse context elements for period information
        contexts = self._parse_contexts(root, nsmap)

        # Parse unit elements
        units = self._parse_units(root, nsmap)

        # Extract facts from all non-structural elements
        for elem in root.iter():
            tag = elem.tag
            if not isinstance(tag, str):
                continue

            # Skip structural elements (contexts, units, footnotes, etc.)
            local_name = etree.QName(tag).localname if "}" in tag else tag
            namespace = etree.QName(tag).namespace if "}" in tag else ""

            # Skip XBRL infrastructure elements
            skip_namespaces = {
                "http://www.xbrl.org/2003/instance",
                "http://www.xbrl.org/2003/linkbase",
                "http://www.w3.org/1999/xlink",
            }
            if namespace in skip_namespaces:
                continue

            # This is a fact element
            text = (elem.text or "").strip()
            if not text:
                continue

            # Build concept name
            concept = self._build_concept_name(namespace, local_name, nsmap)

            # Get context reference for period info
            context_ref = elem.get("contextRef")
            period_info = contexts.get(context_ref, {}) if context_ref else {}

            # Get unit reference
            unit_ref = elem.get("unitRef")
            unit = units.get(unit_ref, self.default_unit) if unit_ref else None

            # Get decimals
            decimals_str = elem.get("decimals")
            decimals = None
            if decimals_str and decimals_str not in ("INF", "inf"):
                try:
                    decimals = int(decimals_str)
                except ValueError:
                    pass

            # Determine value type
            value_numeric = None
            value_text = None
            value_boolean = None

            if text.lower() in ("true", "false"):
                value_boolean = text.lower() == "true"
            else:
                try:
                    value_numeric = float(text.replace(",", ""))
                except ValueError:
                    value_text = text

            # Extract dimension info from context
            dimension_member = period_info.get("dimension_member")
            dimension_value = period_info.get("dimension_value")

            fact = XBRLFact(
                ticker=self.ticker,
                concept=concept,
                label_en=local_name,
                value_numeric=value_numeric,
                value_text=value_text,
                value_boolean=value_boolean,
                unit=unit,
                decimals=decimals,
                period_start=period_info.get("period_start"),
                period_end=period_info.get("period_end"),
                period_instant=period_info.get("period_instant"),
                dimension_member=dimension_member,
                dimension_value=dimension_value,
                source_url=self.source_url,
                filing_id=self.filing_id,
            )
            self.facts.append(fact)

        logger.info("XML parsing complete: %d facts from %s", len(self.facts), file_path.name)
        return self.facts

    def _parse_contexts(self, root, nsmap: dict) -> dict:
        """Parse xbrli:context elements to build period lookup.

        Returns dict mapping context id -> {period_start, period_end, period_instant,
        dimension_member, dimension_value}.
        """
        contexts = {}

        # Try multiple namespace resolution approaches
        context_elements = root.findall(
            ".//{http://www.xbrl.org/2003/instance}context"
        )
        if not context_elements:
            # Fallback: search all elements with 'context' local name
            context_elements = [
                el for el in root.iter()
                if isinstance(el.tag, str) and el.tag.endswith("}context")
            ]

        for ctx in context_elements:
            ctx_id = ctx.get("id")
            if not ctx_id:
                continue

            info = {
                "period_start": None,
                "period_end": None,
                "period_instant": None,
                "dimension_member": None,
                "dimension_value": None,
            }

            # Parse period
            for period_elem in ctx.iter():
                tag = period_elem.tag if isinstance(period_elem.tag, str) else ""
                local = tag.split("}")[-1] if "}" in tag else tag
                text = (period_elem.text or "").strip()

                if local == "startDate" and text:
                    info["period_start"] = self._safe_parse_date(text)
                elif local == "endDate" and text:
                    info["period_end"] = self._safe_parse_date(text)
                elif local == "instant" and text:
                    info["period_instant"] = self._safe_parse_date(text)

            # Parse dimension (scenario/segment explicit members)
            for member_elem in ctx.iter():
                tag = member_elem.tag if isinstance(member_elem.tag, str) else ""
                local = tag.split("}")[-1] if "}" in tag else tag

                if local == "explicitMember":
                    dimension = member_elem.get("dimension", "")
                    value = (member_elem.text or "").strip()
                    if dimension:
                        info["dimension_member"] = dimension
                        info["dimension_value"] = value

            contexts[ctx_id] = info

        return contexts

    def _parse_units(self, root, nsmap: dict) -> dict:
        """Parse xbrli:unit elements to build unit lookup.

        Returns dict mapping unit id -> unit string (e.g., 'SAR').
        """
        units = {}

        unit_elements = root.findall(
            ".//{http://www.xbrl.org/2003/instance}unit"
        )
        if not unit_elements:
            unit_elements = [
                el for el in root.iter()
                if isinstance(el.tag, str) and el.tag.endswith("}unit")
            ]

        for unit_elem in unit_elements:
            unit_id = unit_elem.get("id")
            if not unit_id:
                continue

            # Look for measure element
            for measure in unit_elem.iter():
                tag = measure.tag if isinstance(measure.tag, str) else ""
                local = tag.split("}")[-1] if "}" in tag else tag
                if local == "measure":
                    text = (measure.text or "").strip()
                    # Extract currency code from namespace-prefixed value
                    # e.g., "iso4217:SAR" -> "SAR"
                    if ":" in text:
                        text = text.split(":")[-1]
                    units[unit_id] = text
                    break

        return units

    def _build_concept_name(self, namespace: str, local_name: str, nsmap: dict) -> str:
        """Build a prefixed concept name from namespace and local name.

        Tries to find a matching prefix in the document's namespace map.
        Falls back to known IFRS prefixes.
        """
        # Try to find prefix in document namespace map
        for prefix, ns in nsmap.items():
            if ns == namespace and prefix:
                return f"{prefix}:{local_name}"

        # Check known prefixes
        known = {
            "http://xbrl.ifrs.org/taxonomy": "ifrs-full",
            "http://xbrl.ifrs.org": "ifrs-full",
        }
        for ns_prefix, concept_prefix in known.items():
            if namespace and namespace.startswith(ns_prefix):
                return f"{concept_prefix}:{local_name}"

        # Fallback
        if namespace:
            # Use last path segment as prefix
            parts = namespace.rstrip("/").split("/")
            prefix = parts[-1] if parts else "unknown"
            return f"{prefix}:{local_name}"

        return f"tasi:{local_name}"

    @staticmethod
    def _safe_parse_date(text: str) -> Optional[date]:
        """Parse a date string, returning None on failure."""
        text = text.strip()
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text.split("T")[0] if "T" in text else text, fmt.split("T")[0]).date()
            except ValueError:
                continue
        return None

    # ------------------------------------------------------------------
    # Excel/workbook parsing (openpyxl) - preserved from original
    # ------------------------------------------------------------------

    def process_workbook(self, file_path: Path) -> list[XBRLFact]:
        """Process an Excel workbook and extract XBRL facts.

        Args:
            file_path: Path to the Excel file.

        Returns:
            List of XBRLFact objects extracted from the workbook.
        """
        if openpyxl is None:
            raise ImportError("openpyxl is required: pip install openpyxl")

        if not file_path.exists():
            self.errors.append(f"File not found: {file_path}")
            return []

        self.facts = []
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as e:
            self.errors.append(f"Cannot open workbook {file_path}: {e}")
            return []

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                sheet_facts = self._process_sheet(ws, sheet_name)
                self.facts.extend(sheet_facts)
            except Exception as e:
                self.errors.append(f"Error processing sheet '{sheet_name}': {e}")

        wb.close()
        return self.facts

    def _process_sheet(self, ws, sheet_name: str) -> list[XBRLFact]:
        """Process a single worksheet.

        Expects a tabular layout where:
        - Column A: Concept/line item name (English or Arabic label)
        - Column B+: Period values (headers indicate dates)

        The first row is treated as headers containing period dates.
        """
        facts = []
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            return facts

        # Parse headers to extract period dates
        headers = rows[0]
        period_dates = self._parse_period_headers(headers)

        # Process data rows
        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not row[0]:
                continue

            concept_label = str(row[0]).strip()
            if not concept_label:
                continue

            # Derive concept identifier from label
            concept = self._label_to_concept(concept_label, sheet_name)

            # Detect if label is Arabic
            label_ar = concept_label if self._is_arabic(concept_label) else None
            label_en = concept_label if not self._is_arabic(concept_label) else None

            # Extract values for each period column
            for col_idx in range(1, len(row)):
                if col_idx >= len(headers):
                    break

                cell_value = row[col_idx]
                if cell_value is None:
                    continue

                period_info = period_dates.get(col_idx, {})
                period_end = period_info.get("period_end")
                period_start = period_info.get("period_start")
                period_instant = period_info.get("period_instant")

                # Skip if no period date context at all
                if not period_end and not period_instant:
                    self.errors.append(
                        f"Row {row_idx}, Col {col_idx}: Missing period_date for "
                        f"concept '{concept}' - skipped"
                    )
                    continue

                # Determine value type
                value_numeric = None
                value_text = None
                value_boolean = None

                if isinstance(cell_value, bool):
                    value_boolean = cell_value
                elif isinstance(cell_value, (int, float)):
                    value_numeric = float(cell_value)
                else:
                    text_val = str(cell_value).strip()
                    if text_val:
                        try:
                            value_numeric = float(text_val.replace(",", ""))
                        except ValueError:
                            value_text = text_val

                fact = XBRLFact(
                    ticker=self.ticker,
                    concept=concept,
                    label_en=label_en,
                    label_ar=label_ar,
                    value_numeric=value_numeric,
                    value_text=value_text,
                    value_boolean=value_boolean,
                    unit=self.default_unit if value_numeric is not None else None,
                    period_start=period_start,
                    period_end=period_end,
                    period_instant=period_instant,
                    source_url=self.source_url,
                    filing_id=self.filing_id,
                )
                facts.append(fact)

        return facts

    def _parse_period_headers(self, headers) -> dict:
        """Parse column headers to extract period date information."""
        period_dates = {}
        for idx, header in enumerate(headers):
            if idx == 0 or header is None:
                continue
            header_str = str(header).strip()
            parsed = self._parse_date_string(header_str)
            if parsed:
                period_dates[idx] = parsed
        return period_dates

    def _parse_date_string(self, s: str) -> Optional[dict]:
        """Parse a date string from a header into period context.

        Handles: '2024-12-31', '31/12/2024', 'FY 2024', 'Q1 2024', '2024'
        """
        s = s.strip()

        # ISO format: YYYY-MM-DD and variants
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                d = datetime.strptime(s, fmt).date()
                return {"period_end": d, "period_start": None, "period_instant": None}
            except ValueError:
                continue

        # Fiscal year: FY 2024, FY2024
        s_upper = s.upper().replace(" ", "")
        if s_upper.startswith("FY") and len(s_upper) >= 6:
            try:
                year = int(s_upper[2:6])
                return {
                    "period_end": date(year, 12, 31),
                    "period_start": date(year, 1, 1),
                    "period_instant": None,
                }
            except ValueError:
                pass

        # Quarter: Q1 2024, Q2 2024, etc.
        quarter_ends = {
            "Q1": (3, 31), "Q2": (6, 30), "Q3": (9, 30), "Q4": (12, 31),
        }
        for q, (month, day) in quarter_ends.items():
            if s_upper.startswith(q) and len(s_upper) >= len(q) + 4:
                try:
                    year = int(s_upper[len(q):len(q) + 4])
                    quarter_start_month = month - 2
                    return {
                        "period_end": date(year, month, day),
                        "period_start": date(year, quarter_start_month, 1),
                        "period_instant": None,
                    }
                except ValueError:
                    pass

        # Year only: 2024
        if s.isdigit() and len(s) == 4:
            try:
                year = int(s)
                return {
                    "period_end": date(year, 12, 31),
                    "period_start": date(year, 1, 1),
                    "period_instant": None,
                }
            except ValueError:
                pass

        return None

    def _label_to_concept(self, label: str, sheet_name: str) -> str:
        """Convert a human-readable label to an XBRL-like concept identifier."""
        normalized = label.strip().lower()

        if normalized in self.IFRS_CONCEPT_MAP:
            return self.IFRS_CONCEPT_MAP[normalized]

        # Fallback: create a concept from the label in PascalCase
        words = normalized.replace("-", " ").replace("_", " ").split()
        pascal = "".join(w.capitalize() for w in words if w)

        # Determine prefix based on sheet context
        prefix = "tasi"
        if any(kw in sheet_name.lower() for kw in ["balance", "financial position"]):
            prefix = "ifrs-full"
        elif any(kw in sheet_name.lower() for kw in ["income", "profit", "loss"]):
            prefix = "ifrs-full"
        elif any(kw in sheet_name.lower() for kw in ["cash flow"]):
            prefix = "ifrs-full"

        return f"{prefix}:{pascal}"

    @staticmethod
    def _is_arabic(text: str) -> bool:
        """Check if text contains Arabic characters."""
        for char in text:
            if "\u0600" <= char <= "\u06FF" or "\u0750" <= char <= "\u077F":
                return True
        return False


# ---------------------------------------------------------------------------
# Database operations
# ---------------------------------------------------------------------------

def _get_connection(pg_conn_or_pool):
    """Get a connection, supporting both direct connections and pool context managers."""
    return pg_conn_or_pool


def insert_facts(pg_conn, facts: list[XBRLFact], dry_run: bool = False) -> int:
    """Insert XBRL facts into PostgreSQL. Returns count of rows inserted."""
    if not facts:
        return 0

    if dry_run:
        logger.info("Would insert %d facts (dry run)", len(facts))
        return len(facts)

    rows = [f.to_insert_tuple() for f in facts]
    cur = pg_conn.cursor()
    inserted = 0

    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        psycopg2.extras.execute_batch(cur, XBRL_INSERT_SQL, batch)
        inserted += len(batch)

    pg_conn.commit()
    return inserted


def create_filing(pg_conn, ticker: str, filing_type: str, filing_date: date,
                  source: str, source_url: str, dry_run: bool = False) -> Optional[str]:
    """Create a filing record and return its UUID."""
    if dry_run:
        logger.info("Would create filing: %s %s %s", ticker, filing_type, filing_date)
        return "dry-run-filing-id"

    cur = pg_conn.cursor()
    cur.execute(
        "INSERT INTO filings (ticker, filing_type, filing_date, source, source_url, status) "
        "VALUES (%s, %s, %s, %s, %s, 'processing') RETURNING id",
        (ticker, filing_type, filing_date, source, source_url),
    )
    filing_id = str(cur.fetchone()[0])
    pg_conn.commit()
    return filing_id


def check_filing_exists(pg_conn, ticker: str, source_url: str) -> bool:
    """Check if a filing has already been processed (by ticker + source_url)."""
    cur = pg_conn.cursor()
    cur.execute(
        "SELECT 1 FROM filings WHERE ticker = %s AND source_url = %s AND status = 'completed' LIMIT 1",
        (ticker, source_url),
    )
    return cur.fetchone() is not None


def mark_filing_complete(pg_conn, filing_id: str, dry_run: bool = False) -> None:
    """Mark a filing as completed."""
    if dry_run:
        return
    cur = pg_conn.cursor()
    cur.execute(
        "UPDATE filings SET status = 'completed', processed_at = NOW() WHERE id = %s",
        (filing_id,),
    )
    pg_conn.commit()


def mark_filing_failed(pg_conn, filing_id: str, dry_run: bool = False) -> None:
    """Mark a filing as failed."""
    if dry_run:
        return
    cur = pg_conn.cursor()
    cur.execute(
        "UPDATE filings SET status = 'failed' WHERE id = %s",
        (filing_id,),
    )
    pg_conn.commit()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Process XBRL filings (XML or Excel) into PostgreSQL xbrl_facts table"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file", type=str, help="Path to a single filing (XML or Excel)")
    group.add_argument("--dir", type=str, help="Directory of filings to process")
    group.add_argument("--url", type=str, help="URL of a filing to download and process")

    parser.add_argument("--ticker", type=str, help="Ticker for single file mode (required with --file/--url)")
    parser.add_argument("--ticker-pattern", type=str, default="*",
                        help="Glob pattern for finding files in --dir mode")
    parser.add_argument("--filing-type", type=str, default="annual",
                        choices=["annual", "quarterly", "interim"],
                        help="Filing type (default: annual)")
    parser.add_argument("--source", type=str, default="Tadawul",
                        help="Data source (default: Tadawul)")
    parser.add_argument("--dry-run", action="store_true", help="Print plan without writing")
    parser.add_argument("--pg-host", default=os.environ.get("PG_HOST", "localhost"))
    parser.add_argument("--pg-port", type=int, default=int(os.environ.get("PG_PORT", "5432")))
    parser.add_argument("--pg-dbname", default=os.environ.get("PG_DBNAME", "radai"))
    parser.add_argument("--pg-user", default=os.environ.get("PG_USER", "radai"))
    parser.add_argument("--pg-password", default=os.environ.get("PG_PASSWORD", ""))
    return parser.parse_args()


def process_single_file(
    file_path: Path,
    ticker: str,
    filing_type: str,
    source: str,
    pg_conn,
    dry_run: bool,
) -> tuple[int, list[str]]:
    """Process a single filing file. Returns (facts_count, errors)."""
    print(f"\nProcessing: {file_path.name} (ticker: {ticker})")

    # Skip already-processed filings
    if pg_conn and not dry_run:
        if check_filing_exists(pg_conn, ticker, str(file_path)):
            print(f"  Skipped: already processed")
            return 0, []

    # Create filing record
    filing_date = date.today()
    filing_id = create_filing(
        pg_conn, ticker, filing_type, filing_date,
        source, str(file_path), dry_run,
    )

    # Parse filing
    processor = XBRLProcessor(
        ticker=ticker,
        filing_id=filing_id,
        source_url=str(file_path),
    )
    facts = processor.process_filing(file_path)

    if processor.errors:
        print(f"  Warnings: {len(processor.errors)}")
        for err in processor.errors[:5]:
            print(f"    - {err}")
        if len(processor.errors) > 5:
            print(f"    ... and {len(processor.errors) - 5} more")

    # Insert facts
    count = insert_facts(pg_conn, facts, dry_run)
    print(f"  Facts extracted: {len(facts)}, Inserted: {count}")

    # Update filing status
    if facts:
        mark_filing_complete(pg_conn, filing_id, dry_run)
    else:
        mark_filing_failed(pg_conn, filing_id, dry_run)

    return count, processor.errors


def main():
    args = parse_args()
    t_start = time.time()

    print("=" * 60)
    print("XBRL Processor")
    print("=" * 60)
    if args.dry_run:
        print("MODE: DRY RUN")

    # Connect to PostgreSQL
    pg_conn = None
    if not args.dry_run:
        if psycopg2 is None:
            print("ERROR: psycopg2 is not installed. Install with: pip install psycopg2-binary")
            sys.exit(1)
        try:
            pg_conn = psycopg2.connect(
                host=args.pg_host,
                port=args.pg_port,
                dbname=args.pg_dbname,
                user=args.pg_user,
                password=args.pg_password,
            )
            pg_conn.autocommit = False
        except psycopg2.OperationalError as e:
            print(f"ERROR: Cannot connect to PostgreSQL: {e}")
            sys.exit(1)

    try:
        total_facts = 0
        total_errors = []

        if args.file:
            if not args.ticker:
                print("ERROR: --ticker is required with --file")
                sys.exit(1)

            file_path = Path(args.file)
            count, errors = process_single_file(
                file_path, args.ticker, args.filing_type,
                args.source, pg_conn, args.dry_run,
            )
            total_facts += count
            total_errors.extend(errors)

        elif args.url:
            if not args.ticker:
                print("ERROR: --ticker is required with --url")
                sys.exit(1)

            processor = XBRLProcessor(ticker=args.ticker, source_url=args.url)
            facts = processor.process_url(args.url)
            total_errors.extend(processor.errors)

            if pg_conn and not args.dry_run:
                filing_date = date.today()
                filing_id = create_filing(
                    pg_conn, args.ticker, args.filing_type, filing_date,
                    args.source, args.url, args.dry_run,
                )
                for f in facts:
                    f.filing_id = filing_id
                count = insert_facts(pg_conn, facts, args.dry_run)
                total_facts += count
                if facts:
                    mark_filing_complete(pg_conn, filing_id, args.dry_run)
                else:
                    mark_filing_failed(pg_conn, filing_id, args.dry_run)
            else:
                total_facts += len(facts)

        elif args.dir:
            dir_path = Path(args.dir)
            if not dir_path.exists():
                print(f"ERROR: Directory not found: {dir_path}")
                sys.exit(1)

            files = sorted(dir_path.glob(args.ticker_pattern))
            supported = {".xml", ".xbrl", ".xlsx", ".xls"}
            files = [f for f in files if f.suffix.lower() in supported]
            print(f"Found {len(files)} supported files")

            for file_path in files:
                # Try to extract ticker from filename (e.g., '2222.SR_annual.xlsx')
                ticker = args.ticker
                if not ticker:
                    stem = file_path.stem
                    parts = stem.split("_")
                    if parts and ".SR" in parts[0]:
                        ticker = parts[0]
                    else:
                        print(f"  Skipping {file_path.name}: cannot determine ticker")
                        continue

                count, errors = process_single_file(
                    file_path, ticker, args.filing_type,
                    args.source, pg_conn, args.dry_run,
                )
                total_facts += count
                total_errors.extend(errors)

        # Summary
        elapsed = time.time() - t_start
        print(f"\n{'=' * 60}")
        print(f"Total facts: {total_facts}")
        print(f"Total warnings: {len(total_errors)}")
        print(f"Duration: {elapsed:.1f}s")

    finally:
        if pg_conn is not None:
            pg_conn.close()


if __name__ == "__main__":
    main()
